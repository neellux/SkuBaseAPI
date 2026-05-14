"""Pure reconciliation of DB-truth UPC/alias state against SellerCloud actual state.

Importable by both the dry-run CLI (scripts/daily_sellercloud_sync_dryrun.py) and the
daily poller (services/daily_sellercloud_sync_poller.py). No I/O — the caller hands
us two dicts, we return a deterministic list of Actions.

Action vocabulary (4 types, per plan):
    add_alias     SC is missing a value DB has as a non-primary alias
    delete_alias  SC has an alias DB does not have anywhere on this SKU
    set_primary   SC's BasicInfo UPC differs from DB's primary UPC for this SKU
    clear_primary SC has a primary UPC but DB has no primary for this SKU

Ordering within a SKU (matches execution order in the poller):
    set_primary > clear_primary > add_alias > delete_alias
"""

from __future__ import annotations

import csv
import logging
import sys
from collections import defaultdict

csv.field_size_limit(sys.maxsize)
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class State:
    """Per-SKU UPC state.

    `primary` is the BasicInfo UPC (SC) or the row in child_upcs with
    is_primary_upc=TRUE (DB). May be None.

    `aliases` is the full alias set. By convention this INCLUDES `primary`
    when primary is set. The reconciler canonicalizes inputs to this form.
    """
    primary: str | None
    aliases: frozenset[str]


@dataclass(frozen=True)
class Action:
    sku: str
    value: str            # for clear_primary this is the value being cleared (the old SC primary)
    action: str           # add_alias | delete_alias | set_primary | clear_primary
    db_role: str          # primary | alias | none
    sc_role: str          # primary | alias | none


_ACTION_ORDER = {"set_primary": 0, "clear_primary": 1, "add_alias": 2, "delete_alias": 3}


def _canonicalize(state: State) -> State:
    """Ensure `primary` (when set) is also in `aliases`. Returns a new State."""
    if state.primary and state.primary not in state.aliases:
        return State(state.primary, frozenset(state.aliases | {state.primary}))
    return state


def _role(state: State, value: str) -> str:
    if state.primary == value:
        return "primary"
    if value in state.aliases:
        return "alias"
    return "none"


# ---------------------------------------------------------------------------
# Reconciliation
# ---------------------------------------------------------------------------

@dataclass
class ReconcileResult:
    actions: list[Action] = field(default_factory=list)
    db_conflicts: list[tuple[str, list[str]]] = field(default_factory=list)
    """Values that appear on more than one SKU in DB — surfaced as warnings,
    not actions. `(value, [sku, sku, ...])`."""


def reconcile(
    db_state: dict[str, State],
    sc_state: dict[str, State],
) -> ReconcileResult:
    """Compute the minimal set of actions to make SC match DB.

    Only SKUs present in `db_state` are reconciled. SC aliases on SKUs missing
    from `db_state` (e.g. inactive products) are left alone.
    """
    # Pre-pass: detect DB-internal conflicts (same value on multiple SKUs).
    # We surface these as warnings but the per-SKU diff below still produces
    # correct actions in their presence.
    value_to_skus: dict[str, set[str]] = defaultdict(set)
    for sku, state in db_state.items():
        for v in state.aliases:
            value_to_skus[v].add(sku)
        if state.primary:
            value_to_skus[state.primary].add(sku)
    conflicts = sorted(
        ((v, sorted(skus)) for v, skus in value_to_skus.items() if len(skus) > 1),
        key=lambda t: t[0],
    )

    result = ReconcileResult(db_conflicts=conflicts)

    for sku in sorted(db_state.keys()):
        db = _canonicalize(db_state[sku])
        sc = _canonicalize(sc_state.get(sku, State(None, frozenset())))

        # Primary delta — at most one set_primary or clear_primary per SKU.
        primary_action: Action | None = None
        primary_target_value: str | None = None  # the value being made primary
        primary_cleared_value: str | None = None  # the old SC primary being demoted/cleared

        if db.primary and sc.primary != db.primary:
            primary_action = Action(
                sku=sku,
                value=db.primary,
                action="set_primary",
                db_role="primary",
                sc_role=_role(sc, db.primary),
            )
            primary_target_value = db.primary
            primary_cleared_value = sc.primary  # may be None
        elif not db.primary and sc.primary:
            primary_action = Action(
                sku=sku,
                value=sc.primary,
                action="clear_primary",
                db_role=_role(db, sc.primary),  # likely "none"
                sc_role="primary",
            )
            primary_cleared_value = sc.primary

        if primary_action is not None:
            result.actions.append(primary_action)

        # Alias additions: in db, not in sc.
        # Skip the primary value if we're about to set_primary on it — SC's
        # BasicInfo update will implicitly add it to the alias list, so a
        # separate add_alias would be redundant (and might race).
        for v in sorted(db.aliases - sc.aliases):
            if v == primary_target_value:
                continue
            result.actions.append(Action(
                sku=sku,
                value=v,
                action="add_alias",
                db_role=_role(db, v),
                sc_role="none",
            ))

        # Alias deletions: in sc, not in db.
        # Skip the value we're clearing as primary — set_primary/clear_primary
        # handles it. (If we set a new primary, the old primary value can
        # legitimately stay as an alias in SC IF db also has it as alias,
        # which it doesn't here since it's not in db.aliases — so delete is
        # actually correct. But we let the primary action handle it for
        # cleanliness; subsequent cycles will surface it if SC keeps it.)
        for v in sorted(sc.aliases - db.aliases):
            if v == primary_cleared_value:
                # Already represented by the primary action.
                continue
            result.actions.append(Action(
                sku=sku,
                value=v,
                action="delete_alias",
                db_role="none",
                sc_role=_role(sc, v),
            ))

    return result


def order_per_sku(actions: Iterable[Action]) -> list[Action]:
    """Stable sort: group by SKU (alphabetical), then by action precedence."""
    return sorted(actions, key=lambda a: (a.sku, _ACTION_ORDER[a.action], a.value))


# ---------------------------------------------------------------------------
# TSV loaders (no pandas — csv.reader + dict[str, set[str]])
# ---------------------------------------------------------------------------

def load_sc_aliases(path: str | Path) -> dict[str, set[str]]:
    """Read ExportStandardInfo Kind=2 TSV (header: ProductID\\tAlias)."""
    by_sku: dict[str, set[str]] = defaultdict(set)
    with open(path, "r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f, delimiter="\t")
        header = next(reader, None)
        if header is None:
            return {}
        for row in reader:
            if len(row) < 2:
                continue
            sku, alias = row[0].strip(), row[1].strip()
            if sku and alias:
                by_sku[sku].add(alias)
    return dict(by_sku)


def load_sc_upcs(path: str | Path) -> dict[str, str]:
    """Read ExportCustomInfo TSV (header: ProductID\\tProductName\\tUPC).

    Returns {sku: primary_upc}. Skips rows with empty UPC.
    """
    by_sku: dict[str, str] = {}
    with open(path, "r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f, delimiter="\t")
        header = next(reader, None)
        if header is None:
            return {}
        for row in reader:
            if len(row) < 3:
                continue
            sku, _name, upc = row[0].strip(), row[1], row[2].strip()
            if sku and upc:
                by_sku[sku] = upc
    return by_sku


def build_sc_state(
    aliases_by_sku: dict[str, set[str]],
    primary_by_sku: dict[str, str],
) -> dict[str, State]:
    skus = set(aliases_by_sku) | set(primary_by_sku)
    out: dict[str, State] = {}
    for sku in skus:
        primary = primary_by_sku.get(sku) or None
        aliases = frozenset(aliases_by_sku.get(sku, set()))
        out[sku] = State(primary=primary, aliases=aliases)
    return out


def build_db_state(
    upcs_rows: list[dict],
    keywords_rows: list[dict],
) -> dict[str, State]:
    """Build DB state from two query results.

    upcs_rows: rows of {child_sku, upc, is_primary_upc}
    keywords_rows: rows of {sku, keywords: list[str] | None}
    """
    aliases: dict[str, set[str]] = defaultdict(set)
    primary: dict[str, str] = {}

    for row in upcs_rows:
        sku = row["child_sku"]
        upc = row["upc"]
        if not sku or not upc:
            continue
        aliases[sku].add(upc)
        if row.get("is_primary_upc"):
            primary[sku] = upc

    for row in keywords_rows:
        sku = row["sku"]
        kws = row.get("keywords") or []
        if not sku:
            continue
        for kw in kws:
            if kw:
                aliases[sku].add(kw)

    return {
        sku: State(primary=primary.get(sku), aliases=frozenset(vals))
        for sku, vals in aliases.items()
    }


# ---------------------------------------------------------------------------
# XLSX writer
# ---------------------------------------------------------------------------

def write_actions_xlsx(
    actions: list[Action],
    conflicts: list[tuple[str, list[str]]],
    out_path: str | Path,
) -> None:
    """Write a 2-sheet workbook: `actions` and `summary` (+ `db_conflicts` if any)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()

    ws = wb.active
    ws.title = "actions"
    headers = ["sku", "value", "action", "db_role", "sc_role"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for a in order_per_sku(actions):
        ws.append([a.sku, a.value, a.action, a.db_role, a.sc_role])

    # Summary
    summary_ws = wb.create_sheet("summary")
    summary_ws.append(["action", "count"])
    for cell in summary_ws[1]:
        cell.font = Font(bold=True)
    counts: dict[str, int] = defaultdict(int)
    for a in actions:
        counts[a.action] += 1
    for action_name in sorted(counts):
        summary_ws.append([action_name, counts[action_name]])
    summary_ws.append(["TOTAL", len(actions)])

    if conflicts:
        conflict_ws = wb.create_sheet("db_conflicts")
        conflict_ws.append(["value", "sku_count", "skus"])
        for cell in conflict_ws[1]:
            cell.font = Font(bold=True)
        for value, skus in conflicts:
            conflict_ws.append([value, len(skus), ", ".join(skus)])

    wb.save(str(out_path))


def write_rollback_csv(actions: list[Action], out_path: str | Path) -> None:
    """Inverse-ops CSV: rows describing how to undo each action.

    A `set_primary` to value X can be inverted by `set_primary` back to the
    old SC primary (preserved in `db_role` ... actually we need the old
    sc_primary captured at action time). For now we emit the inverse using
    only what's on Action — caller can enrich.
    """
    with open(out_path, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["sku", "value", "inverse_action", "notes"])
        for a in order_per_sku(actions):
            if a.action == "add_alias":
                w.writerow([a.sku, a.value, "delete_alias", ""])
            elif a.action == "delete_alias":
                w.writerow([a.sku, a.value, "add_alias", ""])
            elif a.action == "set_primary":
                w.writerow([a.sku, a.value, "set_or_clear_primary",
                            "set back to whatever SC primary was before this run"])
            elif a.action == "clear_primary":
                w.writerow([a.sku, a.value, "set_primary",
                            "re-set the cleared value"])
