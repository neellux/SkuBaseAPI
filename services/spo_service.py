import asyncio
import csv

import logging
import os
import re
import tempfile
import traceback
from datetime import datetime
from io import BytesIO
from typing import Any

import httpx
import openpyxl
from tortoise import connections

from config import config
from services.listing_options_service import listing_options_service

logger = logging.getLogger(__name__)

DISPLAY_HEADERS = [
    "Category",
    "SKU",
    "Title",
    "Description",
    "Variant ID",
    "Designer",
    "Image Link 1",
    "Image Link 2",
    "MSRP",
    "Weight",
    "Normalized Color",
    "Unisex Size",
    "Footwear Size",
    "Mens Clothing Bottoms Size",
    "Mens Clothing Tops Size",
    "Womens Clothing Tops Size",
    "Womens Clothing Bottoms Size",
    "Ring Size",
]

API_HEADERS = [
    "category",
    "sku",
    "title",
    "description",
    "variantId",
    "designer",
    "image-link-1",
    "image-link-2",
    "msrp",
    "weight",
    "normalized-color",
    "unisex-size",
    "footwear-size",
    "mens-clothing-bottoms-size",
    "mens-clothing-tops-size",
    "womens-clothing-tops-size",
    "womens-clothing-bottoms-size",
    "ring-size",
]

OFFER_HEADERS = [
    "sku",
    "product-id",
    "product-id-type",
    "price",
    "quantity",
    "state-code",
]

IMAGE_BASE_URL = "https://storage.googleapis.com/lux_products"

TERMINAL_STATUSES = {"COMPLETE", "FAILED", "CANCELLED", "REJECTED"}

MAX_ERROR_DISPLAY_LENGTH = 500

SKU_PATTERN = re.compile(r"^[A-Za-z0-9\-_/. ]+$")


def _sanitize_error_text(text: str) -> str:
    if not text:
        return ""
    clean = re.sub(r"<[^>]+>", "", str(text))
    if len(clean) > MAX_ERROR_DISPLAY_LENGTH:
        clean = clean[:MAX_ERROR_DISPLAY_LENGTH] + "..."
    return clean.strip()


class SpoService:

    PLATFORM_ID = "spo"

    def __init__(self) -> None:
        spo_config = config.get("spo", {})
        self.api_endpoint: str = spo_config.get("api_endpoint", "").rstrip("/")
        self.api_key: str = spo_config.get("api_key", "")
        self._client: httpx.AsyncClient | None = None

    async def _get_client(self) -> httpx.AsyncClient:
        if self._client is None:
            self._client = httpx.AsyncClient(
                timeout=httpx.Timeout(120.0, connect=5.0),
                headers={"Authorization": self.api_key},
            )
        return self._client

    async def close(self) -> None:
        if self._client:
            await self._client.aclose()
            self._client = None

    async def build_product_rows(
        self,
        listing: Any,
        form_data: dict[str, Any],
        field_definitions: list[dict[str, Any]],
    ) -> list[dict[str, Any]]:
        spo_field_map: dict[str, dict[str, Any]] = {}
        for field_def in field_definitions:
            local_name = field_def.get("name")
            if not local_name:
                continue
            for platform in field_def.get("platforms") or []:
                if platform.get("platform_id") == self.PLATFORM_ID:
                    spo_field_map[local_name] = {
                        "field_id": platform.get("field_id"),
                        "platform_tags": platform.get("platform_tags") or [],
                    }
                    break

        row_data: dict[str, Any] = {}
        for local_name, mapping in spo_field_map.items():
            spo_field = mapping["field_id"]
            value = form_data.get(local_name, "")
            row_data[spo_field] = value

        logger.info(f"SPO field map: {list(spo_field_map.keys())}")
        logger.info(f"SPO row_data keys: {list(row_data.keys())}")
        logger.info(f"SPO normalized-color: {row_data.get('normalized-color')}")
        logger.info(f"SPO form standard_color: {form_data.get('standard_color')}")

        category = row_data.get("category")
        if category:
            spo_type = await listing_options_service.get_platform_type(category, self.PLATFORM_ID)
            if spo_type:
                row_data["category"] = spo_type

        weight = row_data.get("weight")
        if weight:
            try:
                row_data["weight"] = round(float(weight) / 16, 1)
            except (ValueError, TypeError):
                row_data["weight"] = 0.0

        child_size_overrides = form_data.get("child_size_overrides", {})
        if not child_size_overrides:
            return []

        sizing_scheme = form_data.get("SIZING_SCHEME", "")
        unique_sizes = list(set(v for v in child_size_overrides.values() if v and str(v).strip()))
        size_map: dict[str, str] = {}
        sizing_type = None
        product_type = form_data.get("product_type")
        if product_type:
            conn = connections.get("default")
            type_result = await conn.execute_query_dict(
                "SELECT sizing_types FROM listingoptions_types WHERE type = $1 LIMIT 1",
                [product_type],
            )
            if type_result:
                sizing_type = type_result[0]["sizing_types"]
        if sizing_scheme and unique_sizes:
            size_map = await listing_options_service.get_mapped_platform_sizes(
                sizing_scheme, unique_sizes, self.PLATFORM_ID, sizing_type
            )
            logger.info(f"SPO size_map: {size_map}")

        pending_pairs: set[tuple[str, str]] = set()
        for mapped_size in size_map.values():
            if mapped_size and " " in mapped_size:
                column, label = mapped_size.split(" ", 1)
                pending_pairs.add((f"{column}-values", label))
        value_code_map = await listing_options_service.get_spo_value_codes(
            list(pending_pairs)
        )

        products = []
        for child_sku, size_str in child_size_overrides.items():
            product = {**row_data}
            product["sku"] = child_sku
            product["variantId"] = listing.product_id
            product["image-link-1"] = f"{IMAGE_BASE_URL}/{listing.product_id}/1_fullsize.jpg"
            product["image-link-2"] = f"{IMAGE_BASE_URL}/{listing.product_id}/2_fullsize.jpg"

            if size_str:
                mapped_size = size_map.get(str(size_str))
                if mapped_size is None:
                    raise ValueError(
                        f"SPO: no platform size mapping for {size_str!r} "
                        f"(child {child_sku}, scheme {sizing_scheme!r}, type {sizing_type!r})"
                    )
                if " " not in mapped_size:
                    raise ValueError(
                        f"SPO: platform_value {mapped_size!r} missing column prefix "
                        f"(child {child_sku}); expected '<column> <label>' format"
                    )
                size_column, size_label = mapped_size.split(" ", 1)
                list_code = f"{size_column}-values"
                value_code = value_code_map.get((list_code, size_label))
                if value_code is None:
                    raise ValueError(
                        f"SPO: no value_code for {list_code}/{size_label!r} "
                        f"(child {child_sku}); config_spo_value_lists is missing this entry"
                    )
                product[size_column] = value_code

            products.append(product)

        return products

    def generate_product_xlsx(self, products: list[dict[str, Any]], output_path: str) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        ws.append(DISPLAY_HEADERS)
        ws.append(API_HEADERS)
        for product in products:
            row = [product.get(field_id) for field_id in API_HEADERS]
            ws.append(row)

        wb.save(output_path)
        logger.info(f"Generated SPO product XLSX: {output_path} ({len(products)} rows)")

    def generate_offer_csv(self, offers: list[dict[str, Any]], output_path: str) -> None:
        with open(output_path, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=OFFER_HEADERS, delimiter=";")
            writer.writeheader()
            for offer in offers:
                writer.writerow(offer)
        logger.info(f"Generated SPO offer CSV: {output_path} ({len(offers)} rows)")

    def build_offer_rows(self, form_data: dict[str, Any]) -> list[dict[str, Any]]:
        child_size_overrides = form_data.get("child_size_overrides", {})
        list_price = form_data.get("list_price", 0)

        offers = []
        for child_sku in child_size_overrides:
            offers.append(
                {
                    "sku": child_sku,
                    "product-id": child_sku,
                    "product-id-type": "SHOP_SKU",
                    "price": list_price,
                    "quantity": 1,
                    "state-code": "11",
                }
            )
        return offers

    async def upload_products(self, xlsx_path: str) -> int:
        client = await self._get_client()
        url = f"{self.api_endpoint}/products/imports"

        with open(xlsx_path, "rb") as f:
            files = {
                "file": (
                    os.path.basename(xlsx_path),
                    f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            }
            response = await client.post(url, files=files, timeout=120.0)

        if response.status_code == 201:
            data = response.json()
            import_id = data.get("import_id")
            logger.info(f"SPO P41 upload successful: import_id={import_id}")
            return import_id
        else:
            raise Exception(f"SPO P41 upload failed: HTTP {response.status_code} - {response.text}")

    async def check_import_status(self, import_id: int) -> dict[str, Any]:
        client = await self._get_client()
        url = f"{self.api_endpoint}/products/imports/{import_id}"
        response = await client.get(url, timeout=30.0)
        response.raise_for_status()
        return response.json()

    async def get_error_report(self, import_id: int) -> list[dict[str, str]]:
        client = await self._get_client()
        url = f"{self.api_endpoint}/products/imports/{import_id}/error_report"
        response = await client.get(url, timeout=60.0)

        if response.status_code == 404:
            return []
        response.raise_for_status()

        errors = []
        content = response.text
        for line in csv.DictReader(content.splitlines(), delimiter=";"):
            sku = line.get("sku", line.get("SKU", ""))
            error_msg = line.get("error-message", line.get("Error Message", ""))
            if sku and SKU_PATTERN.match(sku):
                errors.append(
                    {
                        "sku": sku,
                        "error": _sanitize_error_text(error_msg),
                    }
                )
        return errors

    async def get_transformation_error_report(self, import_id: int) -> list[dict[str, str]]:
        client = await self._get_client()
        url = f"{self.api_endpoint}/products/imports/{import_id}/transformation_error_report"
        response = await client.get(url, timeout=60.0)

        if response.status_code == 404:
            return []
        response.raise_for_status()

        errors = []
        try:
            wb = openpyxl.load_workbook(
                filename=__import__("io").BytesIO(response.content),
                read_only=True,
            )
            ws = wb.active
            headers = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

            sku_idx = None
            errors_idx = None
            for i, h in enumerate(headers):
                h_lower = (h or "").lower().strip()
                if h_lower == "sku":
                    sku_idx = i
                elif h_lower == "errors":
                    errors_idx = i

            if sku_idx is None or errors_idx is None:
                logger.warning(f"P47 report missing sku/errors columns. Headers: {headers}")
                return []

            for row in ws.iter_rows(min_row=2, values_only=True):
                sku = str(row[sku_idx] or "").strip()
                raw_errors = str(row[errors_idx] or "").strip()

                if not sku or not raw_errors:
                    continue
                if not SKU_PATTERN.match(sku):
                    continue

                error_messages = []
                for part in raw_errors.split(","):
                    part = part.strip()
                    if "|" in part:
                        error_messages.append(part.split("|", 1)[1].strip())
                    elif part:
                        error_messages.append(part)

                errors.append(
                    {
                        "sku": sku,
                        "error": _sanitize_error_text("; ".join(error_messages)),
                    }
                )

            wb.close()
        except Exception:
            logger.exception(
                f"Failed to parse P47 transformation error report for import {import_id}"
            )

        return errors

    async def upload_offers(self, csv_path: str) -> int:
        client = await self._get_client()
        url = f"{self.api_endpoint}/offers/imports"

        with open(csv_path, "rb") as f:
            files = {"file": (os.path.basename(csv_path), f, "text/csv")}
            response = await client.post(url, files=files, timeout=120.0)

        if response.status_code == 201:
            data = response.json()
            import_id = data.get("import_id")
            logger.info(f"SPO OF01 upload successful: import_id={import_id}")
            return import_id
        else:
            raise Exception(
                f"SPO OF01 upload failed: HTTP {response.status_code} - {response.text}"
            )

    async def check_offer_status(self, import_id: int) -> dict[str, Any]:
        client = await self._get_client()
        url = f"{self.api_endpoint}/offers/imports/{import_id}"
        response = await client.get(url, timeout=30.0)
        response.raise_for_status()
        return response.json()

    async def get_offer_error_report(self, import_id: int) -> list[dict[str, str]]:
        client = await self._get_client()
        url = f"{self.api_endpoint}/offers/imports/{import_id}/error_report"
        response = await client.get(url, timeout=60.0)

        if response.status_code == 404:
            return []
        response.raise_for_status()

        errors = []
        content = response.text
        for line in csv.DictReader(content.splitlines(), delimiter=";"):
            sku = line.get("sku", line.get("SKU", ""))
            error_msg = line.get("error-message", line.get("Error Message", ""))
            if sku and SKU_PATTERN.match(sku):
                errors.append(
                    {
                        "sku": sku,
                        "error": _sanitize_error_text(error_msg),
                    }
                )
        return errors


spo_service = SpoService()
